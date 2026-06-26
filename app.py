import os
import io
import json
import uuid
import base64
import time
import tempfile
from datetime import datetime

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, session
)
from dotenv import load_dotenv
import anthropic
import openpyxl
import PyPDF2
import re
import markdown
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import httpx
except ImportError:
    httpx = None

try:
    from google import genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")
CLAUDE_MAX_TOKENS = 16384
CLAUDE_MAX_RETRIES = 3
CLAUDE_TIMEOUT_SECONDS = 600.0
REFERENCE_IMAGE_MAX_WIDTH = 600
MAX_BROLL_REFERENCE_IMAGES = 32

load_dotenv(override=True)

app = Flask(__name__)
app.secret_key = "istv-broll-tool-secret-key-2026"
app.config["MAX_CONTENT_LENGTH"] = 4 * 1024 * 1024 * 1024  # 4 GB for large audio uploads

BRIEFS_DIR = os.path.join(os.path.dirname(__file__), ".briefs")
os.makedirs(BRIEFS_DIR, exist_ok=True)

IMAGES_DIR = os.path.join(os.path.dirname(__file__), "generated_images")
os.makedirs(IMAGES_DIR, exist_ok=True)

GENERATED_DIR = os.path.join(os.path.dirname(__file__), "generated")
os.makedirs(GENERATED_DIR, exist_ok=True)

INPUT_DIR = os.path.join(os.path.dirname(__file__), "input")
os.makedirs(INPUT_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {"xlsx", "xls", "pdf", "txt", "mp3", "wav", "m4a", "ogg", "flac", "mp4", "webm"}
AUDIO_EXTENSIONS = {"mp3", "wav", "m4a", "ogg", "flac", "mp4", "webm"}
AUDIO_MAX_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB threshold for compression


def save_to_generated_folder(
    client_full_name, episode_title, brief_md, transcript_text=None, brief_html=None
):
    """Save md/html/docx to generated/<ClientName>_<EpisodeTitle>_<date>/ folder."""
    safe_name = re.sub(r'[^\w\s-]', '', client_full_name or 'Client').strip()
    safe_name = re.sub(r'\s+', '_', safe_name) or 'Client'
    safe_ep = re.sub(r'[^\w\s-]', '', episode_title or 'Episode').strip()
    safe_ep = re.sub(r'\s+', '_', safe_ep) or 'Episode'
    date_str = datetime.now().strftime("%Y-%m-%d_%H%M")

    folder_name = f"{safe_name}_{safe_ep}_{date_str}"
    folder_path = os.path.join(GENERATED_DIR, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    brief_path = os.path.join(folder_path, f"{safe_name}_BRoll_Brief.md")
    with open(brief_path, "w", encoding="utf-8") as f:
        f.write(brief_md)

    if brief_html:
        html_path = os.path.join(folder_path, f"{safe_name}_BRoll_Brief.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(brief_html)
        from brief_document import export_production_brief_docx

        docx_path = os.path.join(folder_path, f"{safe_name}_BRoll_Brief.docx")
        if not export_production_brief_docx(html_path, docx_path, brief_path):
            print(f"Warning: DOCX export failed for {folder_path}")

    if transcript_text:
        transcript_path = os.path.join(folder_path, f"{safe_name}_Transcript.txt")
        with open(transcript_path, "w", encoding="utf-8") as f:
            f.write(transcript_text)

    meta = {
        "client": client_full_name,
        "episode": episode_title,
        "generated_at": datetime.now().isoformat(),
        "files": os.listdir(folder_path),
    }
    meta_path = os.path.join(folder_path, "metadata.json")
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    print(f"Saved to generated folder: {folder_path}")
    return folder_path


def save_brief(brief_html, brief_md, client_full_name, episode_title,
               brief_type="client", images=None, extra_meta=None):
    brief_id = str(uuid.uuid4())
    data = {
        "brief_html": brief_html,
        "brief_md": brief_md,
        "client_full_name": client_full_name,
        "episode_title": episode_title,
        "brief_type": brief_type,
    }
    if extra_meta:
        data["extra_meta"] = extra_meta
    path = os.path.join(BRIEFS_DIR, f"{brief_id}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    if images:
        img_path = os.path.join(BRIEFS_DIR, f"{brief_id}_images.json")
        with open(img_path, "w", encoding="utf-8") as f:
            json.dump(images, f, ensure_ascii=False)
    return brief_id


def load_brief(brief_id):
    path = os.path.join(BRIEFS_DIR, f"{brief_id}.json")
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    img_path = os.path.join(BRIEFS_DIR, f"{brief_id}_images.json")
    if os.path.exists(img_path):
        with open(img_path, "r", encoding="utf-8") as f:
            data["images"] = json.load(f)
    else:
        data["images"] = []
    return data

def _load_master_prompt():
    prompt_path = os.path.join(os.path.dirname(__file__), "prompt_v2.txt")
    with open(prompt_path, "r", encoding="utf-8") as f:
        return f.read()

MASTER_PROMPT = _load_master_prompt()


# ═══════════════════════════════════════════════════════════════
# PROMPT 2 — EDITOR B-ROLL DECISION SHEET (Internal use only)
# ═══════════════════════════════════════════════════════════════

EDITOR_PROMPT = r"""# INSIDE SUCCESS TV — EDITOR B-ROLL DECISION SHEET GENERATOR
## PROMPT 2 (v1.0) — Internal Production Use Only

---

You are a senior post-production supervisor at Inside Success TV. You work exclusively with the editing team — not the client.

The first cut of this episode is complete. Your job is to generate an EDITOR B-ROLL DECISION SHEET — a timestamped, row-by-row reference document that tells the editor exactly what b-roll to use for every segment in the cut, where to source it, and what to do if the primary option is not available.

You will be given a cut sheet. Read every segment carefully. Extract the verbal content (what the subject is saying), the emotional tone, and the visual opportunity each moment presents.

---

## INPUT VARIABLES:

- Subject Name: {subject_name}
- Episode Title: {episode_title}
- Show: {show_name}
- Total Runtime: {total_runtime}
- Editor: {editor_name}
- Stock Library Access: Artgrid, Storyblocks, Pond5, Shutterstock
- AI Generation Tools Available: Midjourney (images), Runway Gen-3, Sora (video)
- Date: {gen_date}

---

## YOUR TASK:

Generate a complete Editor B-Roll Decision Sheet. For EVERY segment in the cut sheet (every row with a timestamp, including VOs), produce a structured entry using the exact format below.

Work through every segment in chronological order. Do not skip any segment, even short VOs.

---

## FORMAT FOR EACH SEGMENT ENTRY:

SEGMENT: [Section name from cut sheet]
TIMESTAMP: [START → END]
TYPE: [IP / VO / TITLE CARD]
VERBAL: [One sentence summary of what is being said — the "verbal" the b-roll must support]
DURATION: [Estimated seconds from cut sheet]
TONE: [Emotional tone from cut sheet]

────────────────────────────────────────
B-ROLL OPTION 1 — CLIENT FOOTAGE [CONFIDENCE: HIGH / MEDIUM / LOW]
What to request: [Specific description of what to ask the client to shoot or pull from archives]
Shot type: [Wide / Medium / Close-up / Archival photo / Archival video]
Why it works: [One sentence — why this shot supports this verbal]
If not available: [Specific alternative client shot to request instead]

────────────────────────────────────────
B-ROLL OPTION 2 — STOCK FOOTAGE [CONFIDENCE: HIGH / MEDIUM / LOW]
Search terms: [3–5 specific search strings to use in Artgrid / Storyblocks / Pond5 / Shutterstock]
  → Primary: "[search term 1]"
  → Alternative 1: "[search term 2]"
  → Alternative 2: "[search term 3]"
What to look for: [Shot description — lighting, movement, mood, subject]
Avoid: [What NOT to pick — e.g. "avoid anything that reads too commercial or staged"]
If nothing fits: [Fallback search term or approach]

────────────────────────────────────────
B-ROLL OPTION 3 — AI GENERATED [CONFIDENCE: HIGH / MEDIUM / LOW]
[ONLY include this section if AI generation is a genuinely good fit for this segment.
Use LOW confidence and skip this section entirely for:
  - Segments about real named locations the subject actually visited
  - Segments featuring the subject's family, children, or specific people
  - Any moment the audience needs to believe is real and documentary
  - Cancer / medical / trauma segments — never use AI here
Use MEDIUM or HIGH confidence only for:
  - Abstract emotional concepts (feeling stuck, survival mode, dreaming)
  - Atmospheric transitions (city at night, open water, landscape)
  - Moments where the visual is suggestive rather than documentary]

AI CONFIDENCE: [HIGH / MEDIUM / LOW — if LOW, write "SKIP — use Client or Stock only" and end this section]

If proceeding:
Midjourney prompt: "[Full Midjourney prompt — cinematic, photorealistic style, specific lighting, mood, composition. Include: --ar 16:9 --style raw --v 6]"
Runway / Sora prompt: "[Full video generation prompt — describe motion, camera movement, duration, mood, lighting]"
What this should look like: [Plain English description of the ideal output]
Best fallback if AI output is unusable: [Specific stock search term OR client footage alternative]

────────────────────────────────────────
EDITOR NOTES:
Duration guidance: [How long this b-roll should run — e.g. "3–5 sec cut", "hold for 8–10 sec"]
Cut style: [How to cut in/out — e.g. "hard cut in, dissolve out", "match cut from previous"]
Priority: [MUST RESOLVE / NICE TO HAVE / STOCK WILL COVER]
Music note: [Any music/score instruction that affects b-roll timing — reference the Music & Score sheet if present]

════════════════════════════════════════

[REPEAT THE ABOVE FORMAT FOR EVERY SEGMENT]

---

## AFTER ALL SEGMENTS — ADD THESE THREE SUMMARY SECTIONS:

---

### SUMMARY A — CLIENT FOOTAGE REQUEST LIST
A consolidated, deduplicated list of everything to request from the client.
Group by category:
- Archival photos needed
- Archival video clips needed
- New lifestyle footage to shoot
List each item once, even if it covers multiple segments.

---

### SUMMARY B — STOCK FOOTAGE PRIORITY LIST
The 10–15 most important stock clips to source, ranked by how critical they are to the edit.
Format: [Priority rank] | [Search term] | [Segment it covers] | [Preferred library]

---

### SUMMARY C — AI GENERATION QUEUE
Only segments where AI generation is rated MEDIUM or HIGH confidence.
Format per item:
- Segment: [name]
- Tool: [Midjourney / Runway / Sora]
- Prompt: [full prompt]
- Fallback: [if unusable]
- Confidence: [MEDIUM / HIGH]

If fewer than 3 segments qualify for AI generation, note: "AI generation minimal for this episode — client and stock footage preferred throughout."

---

## TONE & FORMAT RULES:
- This is an internal editor document. Tone is direct, technical, and efficient.
- No client-friendly language. Write for a professional editor.
- Every stock search term must be specific enough to get useful results — not generic (e.g. NOT "happy woman" — YES "woman walking coastal path golden hour slow motion")
- Every Midjourney prompt must be complete and render-ready — include style, lighting, composition, aspect ratio
- Every Runway/Sora prompt must specify motion, camera movement, and mood
- AI confidence scoring must be honest — do not force AI generation where it does not belong
- Never suggest AI generation for: medical/cancer segments, real family members, real named locations the subject actually visited, or any moment requiring documentary authenticity
- The editor should be able to open this document and immediately know what to do for every single segment

## IMPORTANT OUTPUT FORMAT RULE
Output ONLY the editor decision sheet content in clean Markdown. Do NOT include any preamble, commentary, or explanation before or after the sheet. Start directly with the first segment entry. Use horizontal rules (---) between segments and use bold/headers for structure."""


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_xlsx(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append("\t".join(cells))
    return "\n".join(lines)


def parse_pdf(file_bytes):
    reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"--- Page {i + 1} ---\n{text}")
    return "\n\n".join(pages)


def parse_txt(file_bytes):
    return file_bytes.decode("utf-8", errors="replace")


def parse_file(file_bytes, filename):
    ext = filename.rsplit(".", 1)[1].lower()
    if ext in ("xlsx", "xls"):
        return parse_xlsx(file_bytes)
    elif ext == "pdf":
        return parse_pdf(file_bytes)
    elif ext == "txt":
        return parse_txt(file_bytes)
    raise ValueError(f"Unsupported file type: .{ext}")


def _is_audio_file(filename):
    """Check if file has an audio/video extension for transcription."""
    if not filename or "." not in filename:
        return False
    return filename.rsplit(".", 1)[1].lower() in AUDIO_EXTENSIONS


def _get_rev_api_key():
    """Read the Rev.ai token from .env (REV_AI_TOKEN)."""
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line.startswith("REV_AI_TOKEN="):
                    return line.split("=", 1)[1].strip()
    return os.getenv("REV_AI_TOKEN", "")


def _compress_audio_path(input_path, filename):
    """Compress audio file on disk to mono 16kHz mp3 using ffmpeg. Returns path to compressed file."""
    try:
        import subprocess

        tmp_out_path = input_path + ".compressed.mp3"
        input_size = os.path.getsize(input_path)

        subprocess.run(
            [
                "ffmpeg", "-y", "-i", input_path,
                "-ac", "1", "-ar", "16000", "-b:a", "64k",
                tmp_out_path,
            ],
            capture_output=True,
            timeout=600,
            check=True,
        )

        output_size = os.path.getsize(tmp_out_path)
        print(f"Audio compressed: {input_size/1024/1024:.1f}MB → {output_size/1024/1024:.1f}MB")
        return tmp_out_path
    except Exception as e:
        print(f"Audio compression failed (using original): {e}")
        return input_path


def _transcribe_with_rev_ai(audio_path, filename):
    """Send audio file (on disk) to Rev.ai and return sentence-level transcript text."""
    import requests

    api_key = _get_rev_api_key()
    if not api_key or api_key == "your-rev-key-here":
        raise RuntimeError(
            "REV_AI_TOKEN not set. Add it to your .env file to use audio transcription."
        )

    headers = {"Authorization": f"Bearer {api_key}"}

    file_size = os.path.getsize(audio_path)
    print(f"Uploading {file_size/1024/1024:.1f}MB to Rev.ai...")

    with open(audio_path, "rb") as f:
        submit_resp = requests.post(
            "https://api.rev.ai/speechtotext/v1/jobs",
            headers=headers,
            files={"media": (filename, f, "audio/mpeg")},
            data={"metadata": "istv-broll-tool", "skip_diarization": "false"},
            timeout=300,
        )
    submit_resp.raise_for_status()
    job_id = submit_resp.json()["id"]
    print(f"Rev.ai job submitted: {job_id}")

    for attempt in range(180):
        time.sleep(5)
        status_resp = requests.get(
            f"https://api.rev.ai/speechtotext/v1/jobs/{job_id}",
            headers=headers,
            timeout=30,
        )
        status_resp.raise_for_status()
        status = status_resp.json().get("status")
        if status == "transcribed":
            break
        if status == "failed":
            failure = status_resp.json().get("failure_detail", "Unknown error")
            raise RuntimeError(f"Rev.ai transcription failed: {failure}")
    else:
        raise RuntimeError("Rev.ai transcription timed out after 15 minutes.")

    transcript_resp = requests.get(
        f"https://api.rev.ai/speechtotext/v1/jobs/{job_id}/transcript",
        headers={**headers, "Accept": "text/plain"},
        timeout=30,
    )
    transcript_resp.raise_for_status()
    return transcript_resp.text


def process_audio_file(file_bytes_or_path, filename):
    """Full audio pipeline: save to disk → compress if needed → Rev.ai → return transcript text.
    Accepts either raw bytes or a file path string."""
    if isinstance(file_bytes_or_path, str):
        audio_path = file_bytes_or_path
    else:
        ext = filename.rsplit(".", 1)[1].lower() if "." in filename else "mp3"
        tmp = tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False)
        tmp.write(file_bytes_or_path)
        tmp.close()
        audio_path = tmp.name

    try:
        file_size = os.path.getsize(audio_path)
        if file_size > AUDIO_MAX_SIZE_BYTES:
            print(f"Audio {file_size/1024/1024:.1f}MB exceeds 50MB threshold, compressing...")
            compressed_path = _compress_audio_path(audio_path, filename)
            if compressed_path != audio_path:
                os.unlink(audio_path)
                audio_path = compressed_path

        print(f"Sending {os.path.getsize(audio_path)/1024/1024:.1f}MB to Rev.ai...")
        transcript = _transcribe_with_rev_ai(audio_path, filename)

        if not transcript or not transcript.strip():
            raise RuntimeError("Rev.ai returned empty transcript. Check audio quality.")

        return transcript
    finally:
        if os.path.exists(audio_path):
            os.unlink(audio_path)


def _extract_episode_title_from_cut_sheet(cut_sheet_text: str) -> str:
    """Episode title from cut_sheet banner (after em dash), or 'Episode'."""
    match = re.search(
        r'"id":\s*"cut_sheet"[\s\S]*?"banner_title":\s*"([^"]+)"',
        cut_sheet_text or "",
        re.IGNORECASE,
    )
    if not match:
        match = re.search(r'"banner_title":\s*"([^"]+)"', cut_sheet_text or "")
    if not match:
        return "Episode"
    banner = match.group(1).strip()
    if banner.upper().startswith("APPENDIX"):
        return "Episode"
    if " — " in banner:
        return banner.split(" — ", 1)[1].strip()
    if " - " in banner:
        return banner.split(" - ", 1)[1].strip()
    return banner


def build_prompt_text(client_first_name, client_full_name, episode_title,
                      industry, deadline, editor_notes):
    deadline_display = deadline if deadline.strip() else "TBD"
    editor_notes_display = editor_notes.strip() if editor_notes.strip() else "None provided."

    return MASTER_PROMPT.format(
        client_first_name=client_first_name,
        client_full_name=client_full_name,
        episode_title=episode_title,
        industry=industry,
        deadline=deadline_display,
        deadline_display=deadline_display,
        editor_notes=editor_notes_display,
    )


def build_editor_prompt_text(subject_name, episode_title, show_name,
                             total_runtime, editor_name):
    return EDITOR_PROMPT.format(
        subject_name=subject_name or "[Extract from cut sheet]",
        episode_title=episode_title or "[Extract from cut sheet]",
        show_name=show_name or "Next Level CEO",
        total_runtime=total_runtime or "[Extract from cut sheet]",
        editor_name=editor_name or "TBD",
        gen_date=datetime.now().strftime("%B %d, %Y"),
    )


def get_api_key():
    """Read the key fresh from .env every time to avoid stale cache."""
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line.startswith("ANTHROPIC_API_KEY="):
                    return line.split("=", 1)[1].strip()
    return os.getenv("ANTHROPIC_API_KEY", "")


def _is_retryable_claude_error(exc):
    """True for transient network / stream failures worth retrying."""
    retryable_types = (
        anthropic.APIConnectionError,
        anthropic.APITimeoutError,
        anthropic.InternalServerError,
        anthropic.RateLimitError,
    )
    if isinstance(exc, retryable_types):
        return True
    if httpx is not None and isinstance(
        exc,
        (httpx.RemoteProtocolError, httpx.ReadTimeout, httpx.ConnectError),
    ):
        return True
    msg = str(exc).lower()
    return any(
        phrase in msg
        for phrase in (
            "incomplete chunked read",
            "peer closed connection",
            "connection reset",
            "connection aborted",
            "broken pipe",
            "timed out",
            "timeout",
            "temporarily unavailable",
        )
    )


def _text_from_message(message):
    parts = []
    for block in message.content:
        if getattr(block, "type", None) == "text":
            parts.append(block.text)
    return "".join(parts)


def _anthropic_client(api_key):
    return anthropic.Anthropic(
        api_key=api_key,
        timeout=CLAUDE_TIMEOUT_SECONDS,
        max_retries=0,
    )


def _claude_request_sync(client, content_blocks):
    """Non-streaming request — more reliable for long briefs."""
    message = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=CLAUDE_MAX_TOKENS,
        temperature=0.4,
        messages=[{"role": "user", "content": content_blocks}],
    )
    return _text_from_message(message)


def _run_claude_with_retry(content_blocks):
    """Call Claude with retries and non-streaming-first for stability."""
    api_key = get_api_key()
    if not api_key or api_key == "your-key-here":
        raise RuntimeError(
            "ANTHROPIC_API_KEY not set. Add it to your .env file."
        )

    client = _anthropic_client(api_key)
    last_error = None

    for attempt in range(1, CLAUDE_MAX_RETRIES + 1):
        try:
            return _claude_request_sync(client, content_blocks)
        except Exception as e:
            last_error = e
            if attempt >= CLAUDE_MAX_RETRIES or not _is_retryable_claude_error(e):
                raise
            wait = min(2 ** attempt, 12)
            print(
                f"Claude API attempt {attempt} failed ({e!r}); "
                f"retrying in {wait}s..."
            )
            time.sleep(wait)

    raise last_error


def call_claude(prompt_text, file_bytes, filename):
    """Send the prompt and file directly to Claude. PDFs are sent as native
    document attachments; XLSX/TXT are parsed to text and included inline."""
    api_key = get_api_key()
    if not api_key or api_key == "your-key-here":
        raise RuntimeError(
            "ANTHROPIC_API_KEY not set. Add it to your .env file."
        )

    ext = filename.rsplit(".", 1)[1].lower()
    content_blocks = []

    if ext == "pdf":
        file_b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
        content_blocks.append({
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": file_b64,
            },
        })
        content_blocks.append({
            "type": "text",
            "text": prompt_text,
        })
    else:
        if ext in ("xlsx", "xls"):
            cut_sheet_text = parse_xlsx(file_bytes)
        else:
            cut_sheet_text = parse_txt(file_bytes)

        content_blocks.append({
            "type": "text",
            "text": (
                f"{prompt_text}\n\n"
                f"---\n\n"
                f"## CUT SHEET CONTENT\n\n"
                f"Below is the full text extracted from the client's cut sheet. "
                f"Use this as your sole source of story information.\n\n"
                f"```\n{cut_sheet_text}\n```"
            ),
        })

    return _run_claude_with_retry(content_blocks)


def call_claude_with_text(prompt_text, input_text):
    """Send a prompt with text input (no file) to Claude — used for chaining
    the output of Prompt 1 into Prompt 2."""
    content_blocks = [{
        "type": "text",
        "text": (
            f"{prompt_text}\n\n"
            f"---\n\n"
            f"## CLIENT BRIEF OUTPUT (from Prompt 1)\n\n"
            f"Below is the complete client production brief generated from the "
            f"cut sheet. Use this as your source material to build the editor "
            f"decision sheet.\n\n"
            f"```\n{input_text}\n```"
        ),
    }]

    return _run_claude_with_retry(content_blocks)


def md_to_html(md_text):
    extensions = ["tables", "fenced_code", "nl2br", "sane_lists"]
    return markdown.markdown(md_text, extensions=extensions)


def get_logo_base64():
    for filename in ("logo.svg", "logo.png"):
        logo_path = os.path.join(app.static_folder, filename)
        if os.path.exists(logo_path):
            with open(logo_path, "rb") as f:
                return base64.b64encode(f.read()).decode("utf-8")
    return ""


def get_gemini_api_key():
    """Read the Gemini API key from .env."""
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line.startswith("GEMINI_API_KEY="):
                    return line.split("=", 1)[1].strip()
    return os.getenv("GEMINI_API_KEY", "")


def _normalize_heading_text(line):
    """Strip leading # marks and whitespace."""
    s = line.strip()
    return re.sub(r'^#+\s*', '', s).strip()


def _section3_starts(header_lower):
    """True when this heading begins SECTION 3 (new video / lifestyle b-roll)."""
    if 'photos' in header_lower or 'archival' in header_lower:
        return False
    if 'section 3' in header_lower and any(
        x in header_lower for x in ('video', 'footage', 'lifestyle', 'b-roll', 'broll')
    ):
        return True
    if 'new video footage' in header_lower:
        return True
    hb = header_lower.replace(' ', '')
    if 'lifestyleb-roll' in hb or 'lifestyleb–roll' in hb:
        return True
    if 'current lifestyle' in header_lower and 'shot' in header_lower:
        return True
    return False


def _section3_ends(header_lower):
    """True when this major heading ends SECTION 3."""
    if 'section 4' in header_lower:
        return True
    if 'interview' in header_lower and 'clip' in header_lower:
        return True
    if 'section 5' in header_lower or 'videographer' in header_lower:
        return True
    if 'section 6' in header_lower or (
        'organise' in header_lower and 'submit' in header_lower
    ):
        return True
    if 'section 7' in header_lower or 'submission summary' in header_lower:
        return True
    if 'section 8' in header_lower:
        return True
    return False


INVALID_SHOT_TITLE_SUBSTRINGS = (
    'story-specific',
    'general lifestyle pool',
    'lifestyle pool (shoot',
    'what we need from you',
    'section 3',
    'section 4',
    'easy tier:',
    'medium tier:',
    'tone/mood:',
    'duration:',
)


def _is_valid_shot_title(title):
    t = (title or '').strip()
    if len(t) < 5:
        return False
    tl = t.lower()
    for ph in INVALID_SHOT_TITLE_SUBSTRINGS:
        if ph in tl:
            return False
    if tl in ('shot label', 'duration', 'context', 'easy tier', 'medium tier'):
        return False
    return True


SHOT_LABEL_LINE = re.compile(
    r'^\s*(?:\d+\.\s*)?(?:[-*]\s*)?Shot\s+Label:\s*(.+?)\s*$',
    re.IGNORECASE,
)

NUMBERED_ITEM_LINE = re.compile(r'^\s*\d+\.\s+\S')


def extract_video_shoot_title_from_line(line):
    """Short shot name for Section 3 (numbered Shot Label + legacy formats)."""
    stripped = (line or '').strip()

    m = SHOT_LABEL_LINE.match(stripped)
    if m:
        raw = m.group(1).strip()
        inner = re.match(r'^\*\*(.+?)\*\*\s*$', raw)
        if inner:
            raw = inner.group(1).strip()
        else:
            raw = re.sub(r'^[`"\s]+|[`"\s]+$', '', raw)
        if _is_valid_shot_title(raw):
            return raw
        return None

    m = re.match(
        r'^\d+\.\s*Shot\s+Label:\s*\*\*(.+?)\*\*',
        stripped,
        re.IGNORECASE,
    )
    if m:
        raw = m.group(1).strip()
        if _is_valid_shot_title(raw):
            return raw

    m = re.match(r'^\d+\.\s*\*\*(.+?)\*\*', stripped)
    if m:
        raw = m.group(1).strip()
        if _is_valid_shot_title(raw):
            return raw
        return None

    m = re.match(r'^[-*]\s+\*\*(.+?)\*\*(?:\s*[—–:\-]|$)', stripped)
    if m:
        raw = m.group(1).strip()
        if _is_valid_shot_title(raw):
            return raw
        return None

    m = re.match(r'^\*\*(.+?)\*\*(?:\s*$|\s*[—–:])', stripped)
    if m:
        raw = m.group(1).strip()
        if _is_valid_shot_title(raw):
            return raw
        return None

    m = re.match(r'^\d+\.\s+(.+)$', stripped)
    if m:
        return _title_from_numbered_line(m.group(1).strip())

    return None


SKIP_NUMBERED_LINE_PREFIXES = (
    'transcript context',
    'alternative:',
    'alternative —',
    'minimum requirement',
    'you are ',
    'if more than',
    'if no personal',
    'interview —',
    'interview -',
    'interview:',
)


def _title_from_numbered_line(rest):
    """Parse titles from numbered lines (story shots + lifestyle pool)."""
    rl = rest.lower()
    if any(rl.startswith(p) for p in SKIP_NUMBERED_LINE_PREFIXES):
        return None
    if 'transcript context' in rl:
        return None

    rest = re.sub(r'^Shot\s+Label:\s*', '', rest, flags=re.IGNORECASE).strip()
    bold = re.match(r'^\*\*(.+?)\*\*', rest)
    if bold:
        cand = bold.group(1).strip()
        if _is_valid_shot_title(cand):
            return cand

    short = re.split(r'\s*[—–]\s*', rest)[0].strip()
    short = re.sub(r'^\*+|\*+$', '', short).strip()
    if 8 <= len(short) <= 100 and _is_valid_shot_title(short):
        return short
    return None


def _is_lifestyle_pool_heading(header_lower):
    return any(
        kw in header_lower
        for kw in (
            'general lifestyle',
            'lifestyle pool',
            'section 3b',
            '3b.',
            'b. general',
        )
    )


def extract_locations_for_images(brief_md):
    """Section 3 story-specific shots + General Lifestyle Pool (all video B-roll)."""
    lines = brief_md.split('\n')
    locations = []
    seen_headers = set()
    in_section3 = False
    in_lifestyle_pool = False

    h1_count = sum(
        1 for l in lines
        if l.strip().startswith('# ') and not l.strip().startswith('## ')
    )
    major_level = 1 if h1_count >= 3 else 2

    def _heading_level(line):
        s = line.lstrip()
        lvl = 0
        for ch in s:
            if ch == '#':
                lvl += 1
            else:
                break
        return lvl if lvl and s[lvl:lvl + 1] == ' ' else 0

    def _grab_shot_context(start_idx):
        ctx = []
        for j in range(start_idx + 1, min(start_idx + 18, len(lines))):
            s = lines[j].strip()
            if s.startswith('#'):
                break
            if SHOT_LABEL_LINE.match(s):
                break
            if NUMBERED_ITEM_LINE.match(s) and j > start_idx + 1:
                break
            if re.match(r'^\d+\.\s*\*\*', s) and j > start_idx + 1:
                break
            if s:
                ctx.append(s)
        return ' '.join(ctx[:8])[:450]

    for i, line in enumerate(lines):
        stripped = line.strip()
        lvl = _heading_level(stripped)

        if lvl:
            header_plain = _normalize_heading_text(stripped)
            hl = header_plain.lower()

            if lvl <= major_level + 1 and _section3_starts(hl):
                in_section3 = True
                in_lifestyle_pool = False
                continue

            if in_section3 and lvl <= major_level and _section3_ends(hl):
                in_section3 = False
                in_lifestyle_pool = False
                continue

            if not in_section3:
                continue

            if _is_lifestyle_pool_heading(hl):
                in_lifestyle_pool = True
                continue
            if 'story-specific' in hl:
                in_lifestyle_pool = False
                continue

            title = extract_video_shoot_title_from_line(stripped)
            if title and title.lower() not in seen_headers:
                seen_headers.add(title.lower())
                locations.append({
                    'header': title,
                    'context': _grab_shot_context(i),
                    'pool': in_lifestyle_pool,
                })
            continue

        if not in_section3:
            continue

        title = extract_video_shoot_title_from_line(stripped)
        if title and title.lower() not in seen_headers:
            seen_headers.add(title.lower())
            locations.append({
                'header': title,
                'context': _grab_shot_context(i),
                'pool': in_lifestyle_pool,
            })

    return locations[:MAX_BROLL_REFERENCE_IMAGES]


def safe_client_filename(client_full_name, suffix, ext):
    """Build download filename with client name first for easy filing."""
    safe = re.sub(r'[^\w\s-]', '', client_full_name or 'Client').strip()
    safe = re.sub(r'\s+', '_', safe) or 'Client'
    return f"{safe}_{suffix}.{ext}"


def compress_reference_image(b64, mime):
    """Downscale reference stills for PDF/web (smaller file size, faster load)."""
    try:
        from PIL import Image
        raw = base64.b64decode(b64)
        img = Image.open(io.BytesIO(raw))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        w, h = img.size
        if w > REFERENCE_IMAGE_MAX_WIDTH:
            h = max(1, int(h * REFERENCE_IMAGE_MAX_WIDTH / w))
            w = REFERENCE_IMAGE_MAX_WIDTH
            img = img.resize((w, h), Image.Resampling.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=95, optimize=True)
        return base64.b64encode(buf.getvalue()).decode('utf-8'), 'image/jpeg'
    except Exception as e:
        print(f"Reference image resize skipped: {e}")
        return b64, mime


def build_image_prompt(header, context, lifestyle_pool=False):
    """Create a detailed image generation prompt for Gemini."""
    pool_hint = ""
    if lifestyle_pool:
        pool_hint = (
            "This is a General Lifestyle Pool B-roll shot (evergreen connective "
            "footage for a documentary episode). "
        )
    transcript_hint = ""
    ctx_lower = (context or "").lower()
    if "you just said" in ctx_lower or "transcript" in ctx_lower:
        transcript_hint = (
            " Match the emotional tone implied by the documentary dialogue context. "
        )
    return (
        f"Generate a cinematic, photorealistic photograph for a professional "
        f"TV documentary production brief. {pool_hint}"
        f"Scene: {header}. "
        f"Brief context: {context}. "
        f"{transcript_hint}"
        f"Style: Professional documentary cinematography, natural golden hour "
        f"lighting, shallow depth of field, warm authentic color grade, "
        f"16:9 widescreen composition. The image should feel genuine and "
        f"cinematic, suitable for a broadcast TV documentary reference image. "
        f"No text overlays, no watermarks, no artificial or stock-photo feel."
    )


def build_reference_image_html(img, shot_title=None):
    """HTML block for a B-roll reference still (PDF + Google Docs export)."""
    mime = img.get('mime', 'image/png')
    title = shot_title or img.get('header', 'Reference')
    w = img.get('width', REFERENCE_IMAGE_MAX_WIDTH)
    return (
        f'<div class="broll-reference-still">'
        f'<p class="broll-reference-label">Reference still — {title}</p>'
        f'<img src="data:{mime};base64,{img["b64"]}" '
        f'width="{w}" style="max-width:{w}px;width:100%;height:auto;" '
        f'alt="Reference still for {title}" />'
        f'<p class="broll-reference-caption">Visual reference for videographer — '
        f'not final footage. Shoot to match mood and framing.</p>'
        f'</div>'
    )


MAJOR_SECTION_PAGE_BREAK_KEYWORDS = (
    'what we need from you: photos',
    'photos & archival',
    'photos and archival',
    'what we need from you: new video',
    'new video footage',
    'lifestyle b-roll',
    'what we need from you: interview',
    'interview clips',
    'videographer instructions',
    'how to organise',
    'how to organize',
    'quick submission summary',
)


def _heading_needs_page_break(header_plain):
    """Only top-level content sections get a page break (avoids blank pages)."""
    hl = header_plain.lower().replace('—', '-').replace('–', '-')
    if 'section 1' in hl or 'opening letter' in hl:
        return False
    if 'story-specific' in hl or 'general lifestyle' in hl or 'lifestyle pool' in hl:
        return False
    if hl.startswith('childhood') or hl.startswith('career') or hl.startswith('family'):
        return False
    return any(kw in hl for kw in MAJOR_SECTION_PAGE_BREAK_KEYWORDS)


def _is_major_section_heading(stripped, use_h1):
    if use_h1:
        return (
            stripped.startswith('# ')
            and not stripped.startswith('## ')
        )
    return (
        stripped.startswith('## ')
        and not stripped.startswith('### ')
    )


def generate_location_images(api_key, brief_md):
    """Extract locations from the brief and generate images via Gemini."""
    if not GEMINI_AVAILABLE:
        return []

    from google.genai import types as genai_types

    locations = extract_locations_for_images(brief_md)
    if not locations:
        return []

    client = genai.Client(api_key=api_key)
    results = []

    def _generate_one(loc):
        prompt = build_image_prompt(
            loc['header'],
            loc['context'],
            lifestyle_pool=loc.get('pool', False),
        )
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash-image",
                contents=[prompt],
                config=genai_types.GenerateContentConfig(
                    response_modalities=["IMAGE", "TEXT"],
                ),
            )
            if response.candidates and response.candidates[0].content:
                for part in response.candidates[0].content.parts:
                    if hasattr(part, 'inline_data') and part.inline_data:
                        img_bytes = part.inline_data.data
                        mime = getattr(
                            part.inline_data, 'mime_type', 'image/png'
                        ) or 'image/png'
                        b64 = base64.b64encode(img_bytes).decode('utf-8')
                        b64, mime = compress_reference_image(b64, mime)
                        return {
                            'header': loc['header'],
                            'b64': b64,
                            'mime': mime,
                            'width': REFERENCE_IMAGE_MAX_WIDTH,
                        }
        except Exception as e:
            print(f"Image generation failed for '{loc['header']}': {e}")
        return None

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(_generate_one, loc) for loc in locations]
        for future in as_completed(futures):
            result = future.result()
            if result:
                results.append(result)

    return results


def save_images_to_disk(images_list, client_full_name):
    """Save generated images as files in a client-named folder under
    generated_images/. Returns the folder path."""
    if not images_list:
        return None

    safe_name = re.sub(r'[^\w\s-]', '', client_full_name).strip().replace(' ', '_')
    folder = os.path.join(IMAGES_DIR, safe_name)
    os.makedirs(folder, exist_ok=True)

    for img in images_list:
        header = img['header']
        safe_header = re.sub(r'[^\w\s-]', '', header).strip().replace(' ', '_')
        mime = img.get('mime', 'image/png')
        ext = 'png' if 'png' in mime else 'jpeg'
        filepath = os.path.join(folder, f"{safe_header}.{ext}")
        try:
            with open(filepath, 'wb') as f:
                f.write(base64.b64decode(img['b64']))
        except Exception as e:
            print(f"Failed to save image '{header}': {e}")

    return folder


def _find_matching_image(header, img_lookup):
    """Find a matching image for a header using exact then fuzzy matching."""
    key = header.strip().lower()
    if key in img_lookup:
        return img_lookup[key]
    # Fuzzy: check if most words in an image header appear in this header
    header_words = set(re.sub(r'[^\w\s]', '', key).split())
    for img_key, img in img_lookup.items():
        img_words = set(re.sub(r'[^\w\s]', '', img_key).split())
        if not img_words:
            continue
        overlap = header_words & img_words
        if len(overlap) >= max(1, len(img_words) * 0.6):
            return img
    return None


def build_client_production_brief_html(
    brief_md, client_full_name, episode_title, logo_b64=None
):
    """Client B-Roll letter — updated-7 layout (v2 tables, page breaks, no date line)."""
    from brief_document import build_production_brief

    clean_md, html = build_production_brief(
        brief_md, client_full_name, episode_title, logo_b64=logo_b64
    )
    return clean_md, html


def build_enriched_brief_html(brief_md, images_list=None, embed_images=False):
    """Legacy HTML builder with optional reference images (prefer build_client_production_brief_html)."""
    from brief_document import sanitize_brief_markdown, style_brief_html_tables

    if not embed_images and not images_list:
        fragment_md = sanitize_brief_markdown(brief_md)
        html = md_to_html(fragment_md)
        html = re.sub(
            r"<strong>From your episode:</strong>",
            '<strong class="transcript-label">From your episode:</strong>',
            html,
            flags=re.IGNORECASE,
        )
        html = style_brief_html_tables(html)
        return html

    brief_md = sanitize_brief_markdown(brief_md)
    images_list = images_list or [] if embed_images else []
    lines = brief_md.split('\n')

    img_lookup = {}
    for img in images_list:
        img_lookup[img['header'].strip().lower()] = img

    used_images = set()
    ordered_images = list(images_list)
    ordered_image_idx = 0

    # Determine whether the doc uses # or ## for major sections
    h1_count = sum(
        1 for l in lines
        if l.strip().startswith('# ') and not l.strip().startswith('## ')
    )
    use_h1 = h1_count >= 3

    def _collect_shot_block(start_idx):
        """Collect all lines belonging to one shot (until next shot or header)."""
        block = [lines[start_idx]]
        for j in range(start_idx + 1, len(lines)):
            s = lines[j].strip()
            if s.startswith('#'):
                break
            if SHOT_LABEL_LINE.match(s) and j > start_idx + 1:
                break
            if NUMBERED_ITEM_LINE.match(s) and j > start_idx + 1:
                break
            if re.match(r'^\d+\.\s*\*\*', s) and j > start_idx + 1:
                break
            block.append(lines[j])
        return block

    output = []
    i = 0
    in_video_section = False

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        lvl = 0
        for ch in stripped:
            if ch == '#':
                lvl += 1
            else:
                break
        if lvl and stripped[lvl:lvl + 1] == ' ':
            hl = _normalize_heading_text(stripped).lower()
            if _section3_starts(hl):
                in_video_section = True
            elif in_video_section and _section3_ends(hl):
                in_video_section = False

        shot_title = None
        if in_video_section and not stripped.startswith('#'):
            shot_title = extract_video_shoot_title_from_line(stripped)

        if shot_title:
            matched = _find_matching_image(shot_title, img_lookup)
            if not matched and ordered_image_idx < len(ordered_images):
                candidate = ordered_images[ordered_image_idx]
                if candidate['header'] not in used_images:
                    matched = candidate
            if matched and matched['header'] not in used_images:
                # Output the shot block first, then the image
                block = _collect_shot_block(i)
                output.extend(block)
                used_images.add(matched['header'])
                if matched in ordered_images:
                    ordered_image_idx = ordered_images.index(matched) + 1
                else:
                    ordered_image_idx += 1
                num_match = re.match(r'^(\d+)\.', stripped)
                display_title = (
                    f"{num_match.group(1)}. {shot_title}" if num_match else shot_title
                )
                output.append('')
                output.append(build_reference_image_html(matched, display_title))
                output.append('')
                i += len(block)
                continue

        output.append(line)
        i += 1

    html = md_to_html('\n'.join(output))
    html = re.sub(
        r'<strong>From your episode:</strong>',
        '<strong class="transcript-label">From your episode:</strong>',
        html,
        flags=re.IGNORECASE,
    )
    html = re.sub(r"📌\s*", "", html)
    html = re.sub(r'<div class="section-page-break"></div>\s*', '', html)
    html = _add_major_section_page_break_classes(html)
    html = _wrap_collapsible_sections(html)
    return style_brief_html_tables(html)


def _wrap_collapsible_sections(html):
    """Split HTML into TLDR (always visible) and collapsible detail sections."""
    tldr_pattern = re.compile(
        r'(<h2[^>]*>.*?AT A GLANCE.*?</h2>)(.*?)(?=<h[12][^>]*>)',
        re.DOTALL | re.IGNORECASE,
    )
    tldr_match = tldr_pattern.search(html)
    if not tldr_match:
        return html

    before_tldr = html[:tldr_match.start()]
    tldr_heading = tldr_match.group(1)
    tldr_body = tldr_match.group(2)
    after_tldr = html[tldr_match.end():]

    section_pattern = re.compile(
        r'(<h2[^>]*>)(.*?)(</h2>)(.*?)(?=<h2[^>]*>|$)',
        re.DOTALL,
    )

    sections_html = []
    for m in section_pattern.finditer(after_tldr):
        h2_open = m.group(1)
        h2_text = m.group(2)
        h2_close = m.group(3)
        body = m.group(4)
        plain_text = re.sub(r'<[^>]+>', '', h2_text).strip()
        full_h2 = f'{h2_open}{h2_text}{h2_close}'
        sections_html.append(
            f'<details class="brief-section" open>'
            f'<summary class="brief-section-header">{plain_text}</summary>'
            f'<div class="brief-section-body">{full_h2}{body}</div>'
            f'</details>'
        )

    if not sections_html:
        return html

    result = before_tldr
    result += f'<div class="tldr-box">{tldr_heading}{tldr_body}</div>'
    result += '\n'.join(sections_html)
    return result


def _add_major_section_page_break_classes(html):
    """Page breaks on main section h2 headings — uses both class and inline
    style so page breaks survive clipboard copy into Google Docs."""

    def _replace_h2(match):
        inner = re.sub(r'<[^>]+>', '', match.group(1))
        if _heading_needs_page_break(inner):
            return (
                f'<h2 class="major-section-start" '
                f'style="page-break-before: always;">'
                f'{match.group(1)}</h2>'
            )
        return match.group(0)

    return re.sub(r'<h2>(.*?)</h2>', _replace_h2, html, flags=re.DOTALL)


def build_pdf_html_with_images(brief_md, images_list):
    """Alias for PDF pipeline."""
    return build_enriched_brief_html(brief_md, images_list)


@app.errorhandler(413)
def request_entity_too_large(error):
    flash("File too large. Maximum upload size is 4 GB.", "error")
    return redirect(url_for("index"))


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate():
    """Step 1 — Generate client brief + B-roll images only."""
    file = request.files.get("cut_sheet")
    if not file or file.filename == "":
        flash("Please upload a cut sheet or audio file.", "error")
        return redirect(url_for("index"))

    if not allowed_file(file.filename):
        flash(
            "Invalid file type. Please upload .xlsx, .pdf, .txt, or audio "
            "(.mp3, .wav, .m4a, .ogg, .flac, .mp4, .webm)",
            "error",
        )
        return redirect(url_for("index"))

    is_audio = _is_audio_file(file.filename)
    file_bytes = None
    transcript_text = None

    if is_audio:
        try:
            ext = file.filename.rsplit(".", 1)[1].lower() if "." in file.filename else "mp3"
            tmp = tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False, dir=BRIEFS_DIR)
            file.save(tmp)
            tmp.close()
            audio_path = tmp.name
            print(f"Audio saved to temp: {os.path.getsize(audio_path)/1024/1024:.1f}MB")
            transcript_text = process_audio_file(audio_path, file.filename)
        except Exception as e:
            flash(f"Audio transcription error: {e}", "error")
            return redirect(url_for("index"))
    else:
        try:
            file_bytes = file.read()
        except Exception as e:
            flash(f"Error reading file: {e}", "error")
            return redirect(url_for("index"))

    # --- Collect form fields ---
    client_first_name = request.form.get("client_first_name", "").strip() or "[Extract from cut sheet]"
    client_full_name = request.form.get("client_full_name", "").strip() or "[Extract from cut sheet]"
    episode_title = request.form.get("episode_title", "").strip() or "[Extract from cut sheet]"
    industry = request.form.get("industry", "").strip() or "[Extract from cut sheet]"
    deadline = request.form.get("deadline", "").strip()
    editor_notes = request.form.get("editor_notes", "").strip()

    subject_name = request.form.get("subject_name", "").strip()
    show_name = request.form.get("show_name", "").strip()
    total_runtime = request.form.get("total_runtime", "").strip()
    editor_name = request.form.get("editor_name", "").strip()

    display_name = client_full_name if client_full_name != "[Extract from cut sheet]" else (subject_name or "Client")
    display_title = episode_title if episode_title != "[Extract from cut sheet]" else "Episode"

    # --- STEP 1: Cut sheet → Client Brief ---
    prompt1_text = build_prompt_text(
        client_first_name, client_full_name, episode_title,
        industry, deadline, editor_notes
    )

    try:
        if is_audio and transcript_text:
            client_brief_md = call_claude_with_text(prompt1_text, transcript_text)
        else:
            client_brief_md = call_claude(prompt1_text, file_bytes, file.filename)
    except Exception as e:
        flash(f"Error generating client brief: {e}", "error")
        return redirect(url_for("index"))

    client_images = []
    logo_b64 = get_logo_base64()
    client_brief_md, client_brief_html = build_client_production_brief_html(
        client_brief_md, display_name, display_title, logo_b64=logo_b64
    )

    # Store editor form fields so Step 2 can use them later
    editor_meta = {
        "subject_name": subject_name,
        "show_name": show_name,
        "total_runtime": total_runtime,
        "editor_name": editor_name,
    }

    client_brief_id = save_brief(
        client_brief_html, client_brief_md,
        display_name, display_title, brief_type="client",
        images=client_images, extra_meta=editor_meta,
    )

    save_to_generated_folder(
        display_name, display_title, client_brief_md, transcript_text, client_brief_html
    )

    return render_template(
        "result.html",
        client_brief_html=client_brief_html,
        editor_sheet_html=None,
        client_full_name=display_name,
        episode_title=display_title,
        client_brief_id=client_brief_id,
        editor_brief_id=None,
    )


@app.route("/generate-editor/<client_brief_id>")
def generate_editor(client_brief_id):
    """Step 2 — Generate editor decision sheet from a saved client brief."""
    data = load_brief(client_brief_id)
    if not data:
        flash("Client brief not found. Please generate one first.", "error")
        return redirect(url_for("index"))

    client_brief_md = data["brief_md"]
    client_brief_html = data["brief_html"]
    client_full_name = data["client_full_name"]
    episode_title = data["episode_title"]
    meta = data.get("extra_meta", {})

    prompt2_text = build_editor_prompt_text(
        meta.get("subject_name") or client_full_name,
        episode_title,
        meta.get("show_name"),
        meta.get("total_runtime"),
        meta.get("editor_name"),
    )

    try:
        editor_sheet_md = call_claude_with_text(prompt2_text, client_brief_md)
    except Exception as e:
        flash(f"Error generating editor sheet: {e}", "error")
        return redirect(url_for("index"))

    editor_sheet_html = md_to_html(editor_sheet_md)

    editor_brief_id = save_brief(
        editor_sheet_html, editor_sheet_md,
        client_full_name, episode_title, brief_type="editor"
    )

    return render_template(
        "result.html",
        client_brief_html=client_brief_html,
        editor_sheet_html=editor_sheet_html,
        client_full_name=client_full_name,
        episode_title=episode_title,
        client_brief_id=client_brief_id,
        editor_brief_id=editor_brief_id,
    )


@app.route("/download-pdf/<brief_id>")
def download_pdf(brief_id):
    data = load_brief(brief_id)
    if not data:
        flash("No brief found. Please generate one first.", "error")
        return redirect(url_for("index"))

    brief_html = data["brief_html"]
    brief_md = data.get("brief_md", "")
    images = data.get("images", [])
    client_full_name = data["client_full_name"]
    episode_title = data["episode_title"]
    brief_type = data.get("brief_type", "client")

    if brief_type == "client" and brief_md:
        _, brief_html = build_client_production_brief_html(
            brief_md, client_full_name, episode_title, logo_b64=get_logo_base64()
        )
        pdf_page = brief_html
    else:
        if brief_md:
            brief_html = build_pdf_html_with_images(brief_md, images)
        logo_b64 = get_logo_base64()
        today = datetime.now().strftime("%B %d, %Y")
        pdf_page = render_template(
            "pdf_template.html",
            brief_html=brief_html,
            client_full_name=client_full_name,
            episode_title=episode_title,
            logo_b64=logo_b64,
            generation_date=today,
            brief_type=brief_type,
        )

    try:
        from xhtml2pdf import pisa
        pdf_buffer = io.BytesIO()
        pisa_status = pisa.CreatePDF(io.StringIO(pdf_page), dest=pdf_buffer)
        if pisa_status.err:
            flash("PDF generation failed. Try copying the HTML instead.", "error")
            return redirect(url_for("index"))
        pdf_bytes = pdf_buffer.getvalue()
    except Exception as e:
        flash(f"PDF generation error: {e}", "error")
        return redirect(url_for("index"))

    if brief_type == "editor":
        filename = safe_client_filename(
            client_full_name, "Editor_Decision_Sheet", "pdf"
        )
    else:
        filename = safe_client_filename(
            client_full_name, "Post-Edit_B-Roll_Brief", "pdf"
        )

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/download-docx/<brief_id>")
def download_docx(brief_id):
    """Download Word .docx (updated-7 layout with COM page breaks)."""
    import tempfile

    from brief_document import export_production_brief_docx

    data = load_brief(brief_id)
    if not data:
        flash("No brief found. Please generate one first.", "error")
        return redirect(url_for("index"))

    brief_md = data.get("brief_md", "")
    client_full_name = data["client_full_name"]
    episode_title = data["episode_title"]
    brief_type = data.get("brief_type", "client")

    if brief_type != "client":
        flash("DOCX export is only available for client production briefs.", "error")
        return redirect(url_for("index"))

    if brief_md:
        _, html_doc = build_client_production_brief_html(
            brief_md, client_full_name, episode_title, logo_b64=get_logo_base64()
        )
    else:
        html_doc = data["brief_html"]

    with tempfile.TemporaryDirectory() as tmp:
        html_path = os.path.join(tmp, "brief.html")
        md_path = os.path.join(tmp, "brief.md")
        docx_path = os.path.join(tmp, "brief.docx")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_doc)
        if brief_md:
            with open(md_path, "w", encoding="utf-8") as f:
                f.write(brief_md)
        if not export_production_brief_docx(
            html_path, docx_path, md_path if brief_md else None
        ):
            flash("DOCX export failed. Install Microsoft Word on this machine.", "error")
            return redirect(url_for("index"))
        docx_bytes = open(docx_path, "rb").read()

    filename = safe_client_filename(
        client_full_name, "Post-Edit_B-Roll_Brief", "docx"
    )
    return send_file(
        io.BytesIO(docx_bytes),
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".wordprocessingml.document"
        ),
        as_attachment=True,
        download_name=filename,
    )


@app.route("/download-html/<brief_id>")
def download_html(brief_id):
    """Download a standalone HTML file for copy-paste into Google Docs."""
    data = load_brief(brief_id)
    if not data:
        flash("No brief found. Please generate one first.", "error")
        return redirect(url_for("index"))

    brief_md = data.get("brief_md", "")
    images = data.get("images", [])
    client_full_name = data["client_full_name"]
    episode_title = data["episode_title"]
    brief_type = data.get("brief_type", "client")

    if brief_type == "client":
        if brief_md:
            _, page = build_client_production_brief_html(
                brief_md, client_full_name, episode_title, logo_b64=get_logo_base64()
            )
        else:
            page = data["brief_html"]
    else:
        if brief_md:
            brief_html = build_enriched_brief_html(brief_md)
        else:
            brief_html = data["brief_html"]
        logo_b64 = get_logo_base64()
        today = datetime.now().strftime("%B %d, %Y")
        page = render_template(
            "gdocs_export.html",
            brief_html=brief_html,
            client_full_name=client_full_name,
            episode_title=episode_title,
            logo_b64=logo_b64,
            generation_date=today,
            brief_type=brief_type,
        )

    if brief_type == "editor":
        filename = safe_client_filename(
            client_full_name, "Editor_Decision_Sheet", "html"
        )
    else:
        filename = safe_client_filename(
            client_full_name, "Post-Edit_B-Roll_Brief", "html"
        )

    return send_file(
        io.BytesIO(page.encode("utf-8")),
        mimetype="text/html; charset=utf-8",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/batch", methods=["GET", "POST"])
def batch_generate():
    """Batch-process all cut sheets in the input/ folder.
    GET: show files ready for processing.
    POST: process all files in parallel, return results."""
    input_files = []
    for f in os.listdir(INPUT_DIR):
        ext = f.rsplit(".", 1)[1].lower() if "." in f else ""
        if ext in ALLOWED_EXTENSIONS - AUDIO_EXTENSIONS:
            input_files.append(f)

    if request.method == "GET":
        return render_template("batch.html", input_files=input_files, results=None)

    if not input_files:
        flash("No valid cut sheet files found in the input/ folder. "
              "Add .xlsx, .pdf, or .txt files and try again.", "error")
        return render_template("batch.html", input_files=[], results=None)

    deadline = request.form.get("deadline", "").strip() or "TBD"
    skip_images = request.form.get("skip_images") == "on"

    results = []
    gemini_key = get_gemini_api_key() if not skip_images else None
    gemini_enabled = (
        gemini_key and gemini_key != "your-gemini-key-here"
        and GEMINI_AVAILABLE and not skip_images
    )

    def _process_one(filename):
        filepath = os.path.join(INPUT_DIR, filename)
        try:
            with open(filepath, "rb") as fh:
                file_bytes = fh.read()

            cut_sheet_text = parse_file(file_bytes, filename)
            name_base = filename.rsplit(".", 1)[0].replace("_", " ").replace("-", " ").title()

            prompt_text = build_prompt_text(
                "[Extract from cut sheet]", "[Extract from cut sheet]",
                "[Extract from cut sheet]", "[Extract from cut sheet]",
                deadline, "None provided.",
            )

            ext = filename.rsplit(".", 1)[1].lower()
            if ext == "pdf":
                brief_md = call_claude(prompt_text, file_bytes, filename)
            else:
                brief_md = call_claude_with_text(prompt_text, cut_sheet_text)

            client_images = []
            if gemini_enabled:
                try:
                    client_images = generate_location_images(gemini_key, brief_md)
                except Exception as img_err:
                    print(f"Gemini images failed for {filename}: {img_err}")

            episode_title = _extract_episode_title_from_cut_sheet(cut_sheet_text)
            clean_md, brief_html = build_client_production_brief_html(
                brief_md, name_base, episode_title, logo_b64=get_logo_base64()
            )

            brief_id = save_brief(
                brief_html, clean_md,
                name_base, episode_title,
                brief_type="client", images=client_images,
            )
            save_to_generated_folder(
                name_base, episode_title, clean_md, brief_html=brief_html
            )

            return {
                "filename": filename,
                "success": True,
                "brief_id": brief_id,
                "client_name": name_base,
                "images_count": len(client_images),
            }
        except Exception as e:
            return {
                "filename": filename,
                "success": False,
                "error": str(e),
            }

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            executor.submit(_process_one, f): f for f in input_files
        }
        for future in as_completed(futures):
            results.append(future.result())

    results.sort(key=lambda r: r["filename"])

    succeeded = sum(1 for r in results if r["success"])
    flash(f"Batch complete: {succeeded}/{len(results)} briefs generated.", "success")

    return render_template("batch.html", input_files=input_files, results=results)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8893"))
    app.run(debug=True, host="0.0.0.0", port=port)
